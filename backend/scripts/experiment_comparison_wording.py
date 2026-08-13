"""[실험용, 서비스 미연동] 비교견적서 항목명을 본견적서와 맥락은 같되 다른 단어로 재구성하는
아이디어를 검증하기 위한 독립 스크립트 (PRD 4.3 "다른 단어로 구성" 원칙, 8장 질문 11).

generation_service.py/estimate-wizard.tsx 등 실제 서비스 경로와 완전히 분리되어 있으며,
결과가 괜찮다고 판단되면 그때 서비스에 반영할지 별도로 논의한다. 지금은 로컬에서 결과만 확인한다.

사용법:
    python scripts/experiment_comparison_wording.py --sample
    python scripts/experiment_comparison_wording.py --quote-id <entity_quote_id> --target-entity 블렌디드랩
    python scripts/experiment_comparison_wording.py --all-catalog   # 카탈로그 전체 → xlsx 다운로드
"""

import argparse
import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402

from app.config import CLAUDE_MODEL, get_anthropic, get_supabase  # noqa: E402

SYSTEM_PROMPT = (
    "당신은 대행사 업계에 정통한 영업 담당자입니다. 지금 만드는 것은 '비교견적서'입니다 — 비교견적서는 "
    "본견적서보다 10% 비싸게 책정해서, 고객이 두 견적서를 나란히 놓고 봤을 때 자연스럽게 더 저렴한 "
    "본견적서를 택하게 만드는 용도입니다. 그러려면 비교견적서는 정말로 다른 대행사가 발급한 것처럼 "
    "보여야 합니다 — 같은 서비스를 팔더라도 회사마다 자기 브랜드 언어와 영업 화법으로 다르게 "
    "부르는 것처럼요.\n\n"
    "절대 하지 말 것 ①: 원본 단어를 사전적 동의어로 1:1 치환하는 것(번역기가 돌린 것처럼 어색해짐). "
    "나쁜 예: 'SEO 성능 측정 및 보고' → 'SEO 성과 측정 및 리포팅' (그냥 유의어로 바꾼 것, 다른 "
    "회사가 실제로 쓸 법한 표현이 아님).\n"
    "좋은 예: 'SEO 성능 측정 및 보고' → 'SEO 트래킹 리포트' / '검색 성과 분석 보고' 처럼, 실제 "
    "이 업계 대행사가 자기 상품명으로 붙였을 법한 자연스러운 실무 용어로 바꾸는 것.\n\n"
    "절대 하지 말 것 ② (가장 중요): 핵심 주제어를 완전히 다른 개념으로 바꾸는 것. 항목명에서 "
    "'무엇에 대한 작업인지'를 나타내는 핵심 명사(예: 인터뷰, 설문, 테스트, 모집, 광고, SEO, 데이터, "
    "리포트/보고서, 캠페인 등)는 절대 다른 개념의 단어로 바꾸면 안 됩니다 — 그대로 쓰거나 완전히 "
    "동일한 의미의 표현으로만 바꾸세요. 바꿀 수 있는 건 그 앞뒤 수식어나 동사뿐입니다(작성→제작, "
    "진행→운영 등).\n"
    "나쁜 예: '인터뷰 스크립트 작성' → '토론 가이드 개발' (X, '인터뷰'가 완전히 다른 개념인 "
    "'토론'으로 바뀌어 실제 작업 내용 자체가 달라짐).\n"
    "좋은 예: '인터뷰 스크립트 작성' → '인터뷰 질문지 제작' / 'FGI 스크립트 개발' (O, 핵심 "
    "주제인 인터뷰/FGI는 그대로 유지).\n\n"
    "[기준 법인]이 실제로 쓴 항목명이 주어지면, [대상 법인]이 비슷한 서비스를 자기 스타일대로 "
    "판매한다고 가정하고 항목명을 새로 지으세요. 작업 범위·의미는 동일하게 유지하되(과장/축소 "
    "금지), 표현은 확실히 다르게. 이것은 견적서 상품명이므로 원본처럼 짧고 명료해야 합니다 — "
    "설명문이나 풀어쓴 문장으로 늘어놓지 마세요. module_name(구분)은 바꾸지 마세요 — "
    "item_name(상품명)만 바꿉니다. 반드시 JSON 객체 하나만 응답하세요(설명, 마크다운 코드블록 "
    '없이). 출력 형식: '
    '{"items": [{"module_name": "...", "original": "...", "reworded": "..."}, ...]} '
    "— 입력 항목과 같은 순서, 같은 개수로 응답하세요."
)


def extract_json(text: str) -> dict:
    cleaned = re.sub(r"^```(?:json)?\s*", "", text.strip())
    cleaned = re.sub(r"\s*```$", "", cleaned)
    return json.loads(cleaned)


def reword_items(items: list[dict], target_entity_name: str, reference_entity_name: str) -> list[dict]:
    """items: [{"module_name": str, "item_name": str}, ...] → 같은 순서의 reworded 목록.

    견적서 상품명답게 원본만큼 간결해야 한다는 요구(2026-08-12)를 프롬프트만으로는 못 지킬 때가
    있어, 원본보다 눈에 띄게 길어진(1.3배 초과) 항목이 있으면 그 항목만 짚어 한 번 더 간결하게
    다듬어달라고 재요청한다 — 재요청 문구엔 배수를 언급하지 않는다(사용자 피드백: "몇 배로
    간결하게"가 아니라 "원본처럼 간결해야 한다"는 취지이므로, 내부 트리거로만 쓰고 모델에게는
    질적인 기준으로 전달한다).
    """
    client = get_anthropic()
    item_lines = "\n".join(f'- [{i["module_name"]}] {i["item_name"]}' for i in items)
    user_content = (
        f"[기준 법인(본견적)] {reference_entity_name}\n[대상 법인(비교견적)] {target_entity_name}\n\n"
        f"[본견적서 항목 목록]\n{item_lines}"
    )

    result: list[dict] = []
    for attempt in range(2):
        message = client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=2048,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": user_content}],
        )
        text = next((b.text for b in message.content if b.type == "text"), "")
        data = extract_json(text)
        result = data["items"]
        assert len(result) == len(items), f"입력 {len(items)}개, 응답 {len(result)}개 — 개수가 안 맞습니다."

        too_long = [r for r in result if len(r["reworded"]) > len(r["original"]) * 1.3]
        if not too_long or attempt == 1:
            break
        long_list = "\n".join(f'- "{r["original"]}" → "{r["reworded"]}" (너무 김)' for r in too_long)
        user_content += (
            f"\n\n[재시도] 아래 항목이 견적서 상품명치고 너무 길게 풀어썼습니다. 원본처럼 짧고 "
            f"명료한 상품명 형태로 다시 만드세요(전체 항목 다시 응답):\n{long_list}"
        )
    return result


SAMPLE_ITEMS = [
    {"module_name": "시장검증", "item_name": "FGI 인터뷰이 모집"},
    {"module_name": "시장검증", "item_name": "인터뷰 스크립트 작성"},
    {"module_name": "시장검증", "item_name": "인터뷰 진행"},
    {"module_name": "시장검증", "item_name": "결과 리포트 제공"},
]


def _fetch_all_catalog_items() -> list[dict]:
    """현재(is_current) 카탈로그 전체에서 고유한 (module_name, item_name)만 추린다 — 같은 항목이
    여러 법인에 공유되는 경우(예: 시장검증 4개 표준 모듈)가 많아 중복 실험을 피한다."""
    supabase = get_supabase()
    rows = (
        supabase.table("item_catalogs")
        .select("entity_templates!item_catalogs_entity_id_fkey(name), task_type, module_name, item_name")
        .eq("is_current", True)
        .execute()
        .data
    )
    sources: dict[tuple, set] = {}
    for r in rows:
        key = (r["module_name"], r["item_name"])
        sources.setdefault(key, set()).add(f'{r["entity_templates"]["name"]}/{r["task_type"]}')
    return [
        {"module_name": m, "item_name": i, "sources": ", ".join(sorted(s))}
        for (m, i), s in sorted(sources.items())
    ]


def _chunked(seq: list, size: int) -> list[list]:
    return [seq[i : i + size] for i in range(0, len(seq), size)]


_ENTITY_NAMES = ["테스티파이", "알파브라더스", "블렌디드랩", "썬데이워커", "ABBG"]


def run_all_catalog(output_path: Path) -> None:
    items = _fetch_all_catalog_items()
    print(f"카탈로그 고유 항목 {len(items)}개 발견. 배치로 변환 중...")

    results = []
    for batch_idx, batch in enumerate(_chunked(items, 20)):
        # 모델이 "다른 회사가 실제로 어떻게 부를지" 상상하려면 익명 라벨("본견적 법인")보다
        # 실존하는 법인명을 주는 편이 훨씬 그럴듯한 결과를 낸다(2026-08-12 확인) — 배치마다
        # 서로 다른 실제 법인 쌍을 순환시켜 스타일이 한 방향으로만 굳지 않게 한다.
        reference_entity_name = _ENTITY_NAMES[batch_idx % len(_ENTITY_NAMES)]
        target_entity_name = _ENTITY_NAMES[(batch_idx + 1) % len(_ENTITY_NAMES)]
        reworded = reword_items(batch, target_entity_name=target_entity_name, reference_entity_name=reference_entity_name)
        for src, r in zip(batch, reworded):
            results.append({**r, "sources": src["sources"]})
        print(f"  {len(results)}/{len(items)} 완료")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AI 워딩 변환 실험"

    headers = ["구분(module_name)", "원본 항목명", "변환된 항목명", "카탈로그 출처(법인/과업종류)"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{len(results) + 1}"

    band_fill = PatternFill("solid", fgColor="F3F4F6")
    for i, r in enumerate(results, start=2):
        ws.append([r["module_name"], r["original"], r["reworded"], r["sources"]])
        for col in range(1, 5):
            cell = ws.cell(row=i, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col in (2, 3, 4)))
            if i % 2 == 0:
                cell.fill = band_fill

    # 구분(module_name)이 연속으로 같은 값이면 병합 셀로 묶는다(2026-08-12 요청) —
    # results가 이미 (module_name, item_name) 순으로 정렬돼 있어 같은 값은 항상 붙어 있다.
    merge_start_row = 2
    for idx, r in enumerate(results):
        row = idx + 2
        next_module = results[idx + 1]["module_name"] if idx + 1 < len(results) else None
        if r["module_name"] != next_module:
            if row > merge_start_row:
                ws.merge_cells(start_row=merge_start_row, start_column=1, end_row=row, end_column=1)
                ws.cell(row=merge_start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            merge_start_row = row + 1

    widths = {"A": 18, "B": 32, "C": 32, "D": 30}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 22
    for i in range(2, len(results) + 2):
        ws.row_dimensions[i].height = 30

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"\n완료: {output_path}")


def _fetch_quote_items(quote_id: str) -> tuple[list[dict], str]:
    supabase = get_supabase()
    row = (
        supabase.table("entity_quotes")
        .select("line_items, entity_templates(name)")
        .eq("id", quote_id)
        .execute()
        .data
    )
    if not row:
        raise SystemExit(f"entity_quote를 찾을 수 없습니다: {quote_id}")
    entity_name = row[0]["entity_templates"]["name"]
    items = [{"module_name": li["category"], "item_name": li["name"]} for li in row[0]["line_items"]]
    return items, entity_name


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--sample", action="store_true", help="하드코딩된 샘플 항목으로 실행")
    parser.add_argument("--quote-id", help="실제 entity_quote id (본견적서)로 실행")
    parser.add_argument("--target-entity", default="비교견적 법인", help="비교견적서 대상 법인명")
    parser.add_argument(
        "--all-catalog", action="store_true", help="현재 카탈로그의 모든 고유 항목을 일괄 변환해 xlsx로 저장"
    )
    parser.add_argument(
        "--output",
        default=str(Path(__file__).resolve().parent / "output" / "comparison_wording_experiment.xlsx"),
        help="--all-catalog 결과 저장 경로",
    )
    args = parser.parse_args()

    if args.all_catalog:
        run_all_catalog(Path(args.output))
        return

    if args.quote_id:
        items, reference_entity_name = _fetch_quote_items(args.quote_id)
    else:
        items, reference_entity_name = SAMPLE_ITEMS, "테스티파이"

    reworded = reword_items(items, args.target_entity, reference_entity_name)

    print(f"[기준: {reference_entity_name}] → [대상: {args.target_entity}]\n")
    for r in reworded:
        print(f"- [{r['module_name']}]")
        print(f"    원본  : {r['original']}")
        print(f"    변환  : {r['reworded']}\n")


if __name__ == "__main__":
    main()
