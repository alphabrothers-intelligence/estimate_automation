"""채팅 기반 견적 수정 (PRD 4.4, 7.3) — 2026-08-21 전면 재작성.

실무자는 원래 Claude 채팅창에 견적서를 붙여넣고 대화하며 고쳤다. 그 경험을 앱 안으로 옮기는
게 목표인데, 예전 구현(edit_service)은 정반대로 세팅돼 있어서 품질이 나빴다:

  · 매 요청이 단발 호출 — 대화 이력이 없어 "아까 그거 다시"가 통하지 않았다
  · "설명 없이 JSON만" 강제 — 생각을 글로 쓸 공간을 없애 산수·판단을 즉답시켰다 (품질 저하 1순위)
  · 필드 10개짜리 경직된 스키마 — 모델 역량이 스키마 준수에 소모됐다
  · max_tokens 2048 — 12개 항목이면 잘려서 JSON 파싱 실패 → 502
  · 파일 첨부 불가 — 견적서를 붙여넣을 수가 없었다
  · 답변이 "요청하신 내용을 반영했습니다." 고정 문구 — 대화가 아니라 폼 제출이었다

지금은 **도구 사용(tool use)**으로 바꿨다. 모델은 평소처럼 자연어로 답하고, 표를 고쳐야 할 때만
apply_quote_edit 도구를 부른다. 답변 문장은 채팅창에 그대로 뜨고 도구 입력은 표에 반영된다 —
실제 Claude 채팅과 같은 모양이다.

금액 규칙은 나머지 시스템과 같다: 금액은 언제나 양식 수식으로 다시 계산하고(quote_pricing),
사용자가 목표 총액을 말했을 때만 그 목표에 맞춘다. 말하지 않았으면 고친 항목만 바뀐다.
"""

import base64
import io
from typing import Any, Dict, List, Optional, Tuple

from fastapi import HTTPException

from app.config import CLAUDE_MAX_TOKENS, CLAUDE_MODEL, get_anthropic, get_supabase
from app.models.estimate import EditResult, EntityQuoteOut
from app.services import pdf_service, quote_pricing
from app.services.catalog_service import normalize_description

# 대화가 길어지면 앞부분을 잘라 보낸다 — 견적서 전체가 매 턴 시스템 프롬프트에 들어가므로
# 이력까지 무한정 쌓으면 비용만 늘고 정확도는 오히려 떨어진다.
MAX_HISTORY_TURNS = 20

APPLY_TOOL = {
    "name": "apply_quote_edit",
    "description": (
        "견적서 항목을 고친다. 사용자가 실제로 표를 바꿔달라고 했을 때만 부른다 — "
        "질문에 답하기만 하면 되는 경우에는 부르지 않는다. "
        "고칠 항목만 넣으면 되고, 넣지 않은 항목과 넣지 않은 칸은 그대로 유지된다."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "items": {
                "type": "array",
                "description": "고칠 항목들. 건드리지 않을 항목은 넣지 않는다.",
                "items": {
                    "type": "object",
                    "properties": {
                        "i": {"type": "integer", "description": "항목 번호(견적서에 표시된 1부터의 번호)"},
                        "name": {"type": "string", "description": "새 상품명. 안 바꾸면 생략"},
                        "category": {"type": "string", "description": "새 구분(대). 안 바꾸면 생략"},
                        "mid_category": {
                            "type": "string",
                            "description": "새 구분(중). 알파브라더스·테스티파이처럼 구분이 두 단계인 양식에서만 쓴다. 안 바꾸면 생략",
                        },
                        "description": {
                            "type": "string",
                            "description": "새 상품구성. 세로형 개조식(1. 2. 3.)으로 줄바꿈해서 쓴다. 안 바꾸면 생략",
                        },
                        "work_days": {"type": "integer", "description": "새 작업일(정수). 안 바꾸면 생략"},
                        "quantity": {"type": "integer", "description": "새 수량(정수). 안 바꾸면 생략"},
                        "unit_price": {"type": "integer", "description": "새 단가(원, 만원 단위). 안 바꾸면 생략"},
                    },
                    "required": ["i"],
                },
            },
            "add_items": {
                "type": "array",
                "description": "새로 추가할 항목들. 기존 항목을 고치는 게 아니라 없던 행을 만들 때 쓴다.",
                "items": {
                    "type": "object",
                    "properties": {
                        "after": {
                            "type": "integer",
                            "description": "이 번호의 항목 바로 아래에 넣는다. 0이면 맨 위, 생략하면 맨 아래.",
                        },
                        "name": {"type": "string", "description": "상품명"},
                        "category": {"type": "string", "description": "구분(대). 기존 항목과 같은 묶음이면 그 이름을 그대로 쓴다"},
                        "mid_category": {"type": "string", "description": "구분(중). 두 단계 양식에서만 쓴다"},
                        "description": {
                            "type": "string",
                            "description": "상품구성. 세로형 개조식(1. 2. 3.)으로 줄바꿈해서 쓴다",
                        },
                        "work_days": {"type": "integer", "description": "작업일(정수)"},
                        "quantity": {"type": "integer", "description": "수량(정수)"},
                        "unit_price": {"type": "integer", "description": "단가(원, 만원 단위)"},
                    },
                    "required": ["name", "unit_price"],
                },
            },
            "remove_item_numbers": {
                "type": "array",
                "items": {"type": "integer"},
                "description": "삭제할 항목 번호들. 없으면 생략",
            },
            "target_supply_amount": {
                "type": "integer",
                "description": (
                    "사용자가 전체 공급가액(VAT 별도) 목표를 말한 경우 그 금액. "
                    "'VAT 포함 3000만원'처럼 말했으면 1.1로 나눈 공급가액을 넣는다. "
                    "목표를 말하지 않았으면 생략한다 — 생략하면 고친 항목만 바뀌고 총액은 그만큼 따라 움직인다."
                ),
            },
        },
        "required": [],
    },
}


def _system_prompt(quote: dict, form: quote_pricing.FormSpec, vat_included: bool) -> str:
    items = quote.get("line_items") or []
    lines = []
    for n, it in enumerate(items, 1):
        parts = [f"{n}. [{it.get('category') or '-'}] {it.get('name')}"]
        if it.get("work_days") is not None:
            parts.append(f"작업일 {it['work_days']}")
        if it.get("quantity") is not None:
            parts.append(f"수량 {it['quantity']}")
        parts.append(f"단가 {int(it.get('unit_price') or 0):,}원")
        parts.append(f"금액 {int(it.get('amount') or 0):,}원")
        block = " | ".join(parts)
        if it.get("description"):
            block += "\n   상품구성: " + it["description"].replace("\n", " / ")
        lines.append(block)
    supply = sum(i.get("amount") or 0 for i in items)

    return f"""당신은 견적서 작성을 돕는 담당자입니다. 사용자와 대화하며 아래 견적서를 함께 고칩니다.

[이 견적서]
- 발행 법인: {quote['entity_templates']['name']} ({'본견적서' if quote['is_primary'] else '비교견적서'})
- 과업종류: {'/'.join(quote.get('task_types') or [])}
- 공급가액 소계: {supply:,}원 / VAT {'포함' if vat_included else '별도'} 총액: {quote_pricing.grand_total(supply, vat_included):,}원

[이 법인 양식의 규칙]
- {form.formula_text}
- 단가는 {form.unit_price_unit:,}원 단위로 떨어지는 게 보통입니다.
- 작업일·수량은 "며칠", "몇 회/몇 인"이라 정수입니다.

[현재 항목]
{chr(10).join(lines) if lines else "(항목 없음)"}

[일하는 방식]
- 사람에게 말하듯 답합니다. 무엇을 어떻게 바꿨는지, 왜 그렇게 했는지 한두 문장으로 설명하세요.
- 표를 실제로 고쳐야 할 때만 apply_quote_edit 도구를 부릅니다. 질문에 답하기만 하면 되는 경우
  (예: "이 항목이 왜 이 금액이죠?")에는 도구를 부르지 말고 말로 답하세요.
- 금액 계산은 서버가 양식 수식으로 다시 맞춥니다. 단가만 정확히 정하면 되고, 합계를 손으로
  맞추려고 애쓰지 마세요.
- 있는 항목을 고칠 때는 items에 번호(i)로, 없던 행을 새로 만들 때는 add_items에 넣으세요.
  없는 번호를 items에 쓰면 안 됩니다. 지울 때는 remove_item_numbers입니다.
- 사용자가 전체 목표 금액을 말했으면 target_supply_amount에 넣으세요. 말하지 않았으면 넣지
  않습니다 — 그러면 고친 항목만 바뀌고 나머지는 그대로 있습니다.
- 애매하면 임의로 정하지 말고 되물으세요.
- 첨부 파일이 있으면 그 내용을 근거로 판단하세요."""


def _xlsx_to_text(data: bytes, filename: str) -> str:
    """xlsx를 표 형태 텍스트로 바꾼다.

    Anthropic API는 PDF·이미지만 문서 블록으로 받고 xlsx는 못 읽는다. 실무자가 견적서를
    엑셀로 들고 있는 경우가 많아 서버에서 텍스트로 펴서 넣는다(2026-08-21).
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    out = [f"[첨부 파일: {filename}]"]
    for ws in wb.worksheets:
        out.append(f"\n## 시트: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = ["" if c is None else str(c).strip() for c in row]
            if any(cells):
                out.append(" | ".join(cells).rstrip(" |"))
    return "\n".join(out)


def _attachment_blocks(attachment: Optional[Dict[str, str]]) -> List[dict]:
    """첨부 파일을 Claude가 읽을 수 있는 content 블록으로 바꾼다."""
    if not attachment:
        return []
    name = attachment.get("filename") or "첨부파일"
    try:
        raw = base64.b64decode(attachment["data"])
    except Exception as e:  # noqa: BLE001 — 잘못된 base64는 사용자에게 그대로 알린다
        raise HTTPException(status_code=400, detail=f"첨부 파일을 읽지 못했습니다: {e}") from e

    if name.lower().endswith(".pdf"):
        # PDF는 API가 직접 읽는다(document 블록) — 표·레이아웃까지 그대로 본다.
        return [{
            "type": "document",
            "source": {"type": "base64", "media_type": "application/pdf", "data": attachment["data"]},
            "title": name,
        }]
    if name.lower().endswith((".xlsx", ".xlsm")):
        return [{"type": "text", "text": _xlsx_to_text(raw, name)}]
    raise HTTPException(status_code=400, detail="PDF 또는 xlsx 파일만 첨부할 수 있습니다.")


def _apply(items: List[dict], edit: dict, form: quote_pricing.FormSpec) -> Tuple[List[dict], Optional[int]]:
    """도구 입력을 현재 항목에 적용한다. 언급하지 않은 항목·칸은 그대로 둔다."""
    result = [dict(it) for it in items]
    fields = ("name", "category", "mid_category", "description", "work_days", "quantity", "unit_price")

    # 먼저 수정. 삭제·추가로 번호가 밀리기 전에 적용해야 사용자가 본 번호와 맞는다.
    for change in edit.get("items") or []:
        idx = int(change.get("i", 0)) - 1
        if not 0 <= idx < len(result):
            # 없는 번호를 고치라는 건 대개 "추가"를 잘못된 칸에 쓴 것이다. 상품명이 있으면
            # 추가로 받아준다 — 예전엔 여기서 조용히 버려서, 모델은 추가했다고 답하는데 표에는
            # 아무 일도 없었다(2026-08-21 사용자 신고).
            if change.get("name"):
                edit.setdefault("add_items", []).append(
                    {k: v for k, v in change.items() if k != "i"}
                )
            continue
        for key in fields:
            if change.get(key) is not None:
                result[idx][key] = change[key]
        if change.get("description"):
            result[idx]["description"] = normalize_description(change["description"])

    remove = {int(n) - 1 for n in (edit.get("remove_item_numbers") or [])}
    if remove:
        result = [it for n, it in enumerate(result) if n not in remove]

    # 추가는 마지막이다. after는 삭제 전 번호 기준이라 이름으로 다시 찾아 그 뒤에 끼운다 —
    # 번호로만 넣으면 같은 요청에서 앞 항목을 지웠을 때 엉뚱한 자리에 들어간다.
    for new_item in edit.get("add_items") or []:
        row = {
            "name": new_item.get("name") or "신규 항목",
            "category": new_item.get("category"),
            "mid_category": new_item.get("mid_category"),
            "description": normalize_description(new_item.get("description") or ""),
            "work_days": new_item.get("work_days") or 1,
            "quantity": new_item.get("quantity") or 1,
            "unit_price": int(new_item.get("unit_price") or 0),
            "amount": 0,  # finalize가 양식 수식으로 다시 계산한다
        }
        after = new_item.get("after")
        pos = len(result)
        if after is not None:
            anchor_idx = int(after) - 1
            anchor = items[anchor_idx].get("name") if 0 <= anchor_idx < len(items) else None
            if anchor:
                pos = next(
                    (n + 1 for n, it in enumerate(result) if it.get("name") == anchor), len(result)
                )
            elif int(after) == 0:
                pos = 0
        result.insert(pos, row)

    target = edit.get("target_supply_amount")
    return result, int(target) if target else None


def edit_entity_quote(
    entity_quote_id: str, edit_request_text: str, attachment: Optional[Dict[str, str]] = None
) -> EditResult:
    supabase = get_supabase()
    quote_res = (
        supabase.table("entity_quotes")
        .select(
            "id, entity_id, is_primary, task_type, task_types, line_items, estimate_set_id, selected_modules, "
            "chat_history, adjustment_note, markup_ratio, is_catalog_borrowed, catalog_source_entity_name, "
            "service_name, quote_date, recipient_name, recipient_contact, recipient_phone, recipient_email, "
            "entity_templates(name)"
        )
        .eq("id", entity_quote_id)
        .execute()
    )
    if not quote_res.data:
        raise HTTPException(status_code=404, detail="견적서를 찾을 수 없습니다.")
    quote = quote_res.data[0]
    vat_included = (
        supabase.table("estimate_sets").select("vat_included").eq("id", quote["estimate_set_id"]).execute()
    ).data[0]["vat_included"]
    form = pdf_service.resolve_form_spec(
        supabase, quote["entity_id"], quote["task_types"], quote.get("selected_modules")
    )

    history: List[dict] = (quote.get("chat_history") or [])[-MAX_HISTORY_TURNS * 2:]
    user_content: List[Any] = _attachment_blocks(attachment) + [{"type": "text", "text": edit_request_text}]
    messages = history + [{"role": "user", "content": user_content}]

    message = get_anthropic().messages.create(
        model=CLAUDE_MODEL,
        max_tokens=CLAUDE_MAX_TOKENS,
        # 견적서 전체가 들어가는 시스템 프롬프트라 매 턴 다시 처리하면 비싸다 — 캐시에 태운다.
        system=[{
            "type": "text",
            "text": _system_prompt(quote, form, vat_included),
            "cache_control": {"type": "ephemeral"},
        }],
        tools=[APPLY_TOOL],
        messages=messages,
    )

    reply = "\n".join(b.text for b in message.content if b.type == "text").strip()
    edit = next((b.input for b in message.content if b.type == "tool_use"), None)

    items = quote.get("line_items") or []
    changed: List[dict] = []
    if edit:
        items, target = _apply(items, edit, form)
        items, residual, log = quote_pricing.finalize(items, target, form)

        # 모델이 생각만 하고 도구를 바로 부르면 text 블록이 없다(thinking이 기본 활성이라 흔하다).
        # 예전엔 그때 "무엇을 고칠지 조금 더 알려주세요"가 떴다 — 표는 멀쩡히 고쳐졌는데 못
        # 알아들은 척하는 답이라 사용자가 같은 요청을 되풀이하게 만들었다(2026-08-21 사용자 신고).
        if not reply:
            done = []
            if edit.get("add_items"):
                done.append(f"{len(edit['add_items'])}개 항목을 추가")
            if edit.get("remove_item_numbers"):
                done.append(f"{len(edit['remove_item_numbers'])}개 항목을 삭제")
            if edit.get("items"):
                done.append(f"{len(edit['items'])}개 항목을 수정")
            reply = ("요청하신 대로 " + ", ".join(done) + "했습니다.") if done else "표에 반영했습니다."

        # 금액 후처리 결과는 문장 다음에 붙인다 — 위 요약보다 먼저 붙이면 reply가 비지 않게 돼
        # 요약이 통째로 빠진다.
        if residual:
            reply += f"\n\n(격자상 목표에 {residual:+,}원 못 맞췄습니다. 화면에서 직접 고치실 수 있습니다.)"
        elif log:
            reply += "\n\n" + " / ".join(log)

        # 추가된 행도 "방금 바뀐 항목"으로 표시해야 화면에서 어디가 늘었는지 보인다.
        touched = {int(c["i"]) - 1 for c in (edit.get("items") or []) if c.get("i")}
        added_names = {a.get("name") for a in (edit.get("add_items") or []) if a.get("name")}
        changed = [
            it for n, it in enumerate(items) if n in touched or it.get("name") in added_names
        ]

    # 이력에는 이번 턴의 사용자 발화와 모델 응답을 남긴다. 첨부 파일 원본은 다시 보내면
    # 매 턴 비용이 붙으므로 파일명만 남기고 뺀다.
    trimmed_user = [b for b in user_content if b.get("type") == "text"]
    if attachment:
        trimmed_user.insert(0, {"type": "text", "text": f"[첨부: {attachment.get('filename')}]"})
    supabase.table("entity_quotes").update({
        "chat_history": history + [
            {"role": "user", "content": trimmed_user},
            {"role": "assistant", "content": [{"type": "text", "text": reply or "(응답 없음)"}]},
        ]
    }).eq("id", entity_quote_id).execute()

    return EditResult(
        scope="quote_only",
        reply=reply or "무엇을 고칠지 조금 더 알려주세요.",
        entity_quote=EntityQuoteOut(
            id=quote["id"],
            entity_id=quote["entity_id"],
            entity_name=quote["entity_templates"]["name"],
            is_primary=quote["is_primary"],
            task_type=quote["task_type"],
            task_types=quote["task_types"],
            total_amount=quote_pricing.grand_total(sum(i["amount"] for i in items), vat_included),
            line_items=items,
            is_catalog_borrowed=quote["is_catalog_borrowed"],
            catalog_source_entity_name=quote["catalog_source_entity_name"],
            service_name=quote["service_name"],
            quote_date=quote["quote_date"],
            recipient_name=quote["recipient_name"],
            recipient_contact=quote["recipient_contact"],
            recipient_phone=quote["recipient_phone"],
            recipient_email=quote["recipient_email"],
            adjustment_note=quote.get("adjustment_note"),
            markup_ratio=float(quote["markup_ratio"]) if quote.get("markup_ratio") is not None else None,
            **pdf_service.get_column_display(
                supabase, quote["entity_id"], quote["task_types"], quote.get("selected_modules")
            ),
        ),
        changed_items=changed,
    )
